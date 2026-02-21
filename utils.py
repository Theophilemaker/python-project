import hashlib
import os
import random
import string
from datetime import datetime, date, timedelta
from PIL import Image
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pandas as pd
import io

def generate_sku():
    """Generate a random SKU"""
    prefix = ''.join(random.choices(string.ascii_uppercase, k=2))
    numbers = ''.join(random.choices(string.digits, k=6))
    return f"{prefix}{numbers}"

def generate_transfer_number():
    """Generate a unique transfer number"""
    date_str = datetime.now().strftime('%Y%m%d')
    random_str = ''.join(random.choices(string.digits, k=3))
    return f"TR{date_str}{random_str}"

def generate_po_number():
    """Generate a unique purchase order number"""
    date_str = datetime.now().strftime('%Y%m%d')
    random_str = ''.join(random.choices(string.digits, k=3))
    return f"PO{date_str}{random_str}"

def generate_expense_number():
    """Generate a unique expense number"""
    date_str = datetime.now().strftime('%Y%m%d')
    random_str = ''.join(random.choices(string.digits, k=4))
    return f"EXP{date_str}{random_str}"

def calculate_tax(profit, tax_rate=18.00):
    """Calculate tax amount based on profit"""
    return profit * (tax_rate / 100)

def calculate_profit(selling_price, buying_price, quantity=1):
    """Calculate total profit"""
    return (selling_price - buying_price) * quantity

def format_currency(amount):
    """Format amount as currency"""
    return f"{amount:,.0f} RWF"

def save_uploaded_file(file, folder, filename=None):
    """Save uploaded file and return path"""
    if not filename:
        filename = file.filename
    filepath = os.path.join(folder, filename)
    file.save(filepath)
    return filepath

def resize_image(input_path, output_path, size=(200, 200)):
    """Resize an image"""
    img = Image.open(input_path)
    img.thumbnail(size)
    img.save(output_path)

def log_audit(db, user_id, action, details=''):
    """Log an audit trail entry"""
    from models import AuditLog
    import flask
    audit = AuditLog(
        user_id=user_id,
        action=action,
        details=details,
        ip_address=flask.request.remote_addr,
        user_agent=flask.request.user_agent.string
    )
    db.session.add(audit)
    db.session.commit()

def generate_sales_report_pdf(data, start_date, end_date, company_info):
    """Generate PDF sales report"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=72, leftMargin=72, topMargin=72, bottomMargin=72)
    styles = getSampleStyleSheet()
    elements = []
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#4361ee'),
        alignment=1,  # Center alignment
        spaceAfter=30
    )
    elements.append(Paragraph(f"{company_info['name']}", title_style))
    elements.append(Paragraph("Sales Report", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    # Date range
    date_text = f"Period: {start_date.strftime('%d/%m/%Y')} - {end_date.strftime('%d/%m/%Y')}"
    elements.append(Paragraph(date_text, styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Summary table
    summary_data = [
        ['Total Sales', f"{data['total_sales']:,.0f} RWF"],
        ['Total Transactions', str(data['total_transactions'])],
        ['Average Sale', f"{data['average_sale']:,.0f} RWF"],
        ['Active Cashiers', str(data['active_cashiers'])]
    ]
    
    summary_table = Table(summary_data, colWidths=[200, 200])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4361ee')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8f9fa')),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 20))
    
    # Daily breakdown
    if data['daily_sales']:
        elements.append(Paragraph("Daily Breakdown", styles['Heading3']))
        daily_data = [['Date', 'Transactions', 'Sales']]
        for day in data['daily_sales']:
            daily_data.append([
                day['date'].strftime('%d/%m/%Y'),
                str(day['transactions']),
                f"{day['total']:,.0f} RWF"
            ])
        
        daily_table = Table(daily_data, colWidths=[150, 100, 200])
        daily_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4361ee')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(daily_table)
    
    # Build PDF
    doc.build(elements)
    buffer.seek(0)
    return buffer

def generate_excel_report(data, filename):
    """Generate Excel report"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"
    
    # Headers
    headers = ['Date', 'Product', 'Quantity', 'Total', 'Payment', 'Cashier', 'Branch']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4361ee", end_color="4361ee", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data
    for row, sale in enumerate(data, 2):
        ws.cell(row=row, column=1, value=sale['sale_date'].strftime('%d/%m/%Y'))
        ws.cell(row=row, column=2, value=sale['product_name'])
        ws.cell(row=row, column=3, value=sale['quantity'])
        ws.cell(row=row, column=4, value=float(sale['total_price']))
        ws.cell(row=row, column=5, value=sale['payment_method'])
        ws.cell(row=row, column=6, value=sale['cashier_name'])
        ws.cell(row=row, column=7, value=sale['branch_name'])
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_letter].width = min(adjusted_width, 30)
    
    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer